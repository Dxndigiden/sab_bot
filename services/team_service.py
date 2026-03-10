"""
Бизнес-логика бота — между хендлерами и CRUD.
Хендлеры только принимают апдейты и вызывают сервисы.
"""

from aiogram import Bot

from core.config import settings
from database import crud
from schemas.team import TeamCreate, TeamRead


async def register_team(data: TeamCreate) -> TeamRead | None:
    """
    Регистрирует команду. None — дубликат.
    Если MAX_TEAMS задан и лимит достигнут — тоже None (но отдельный кейс — ниже).
    """
    if settings.MAX_TEAMS and (await crud.get_stats())['total'] >= settings.MAX_TEAMS:
        raise RegistrationClosedError('Достигнут лимит команд')

    team = await crud.create_team(data)
    if team is None:
        return None
    return TeamRead.model_validate(team)


class RegistrationClosedError(Exception):
    pass


async def send_confirmation_requests(bot: Bot) -> int:
    """
    Рассылает запросы подтверждения только тем, кто ещё не подтвердил.
    Сохраняет confirm_msg_id, чтобы потом удалить сообщение.
    Возвращает количество успешно отправленных.
    """
    from bot.keyboards.inline import confirm_keyboard

    teams = await crud.get_unconfirmed_teams()
    sent = 0
    for team in teams:
        try:
            msg = await bot.send_message(
                chat_id=team.telegram_id,
                text=(
                    '🏀 <b>SAB Tournament</b>\n\n'
                    'Подтвердите участие вашей команды <b>{}</b> в турнире.\n'
                    'Если не подтвердите — заявка будет отклонена.'
                ).format(team.team_name),
                reply_markup=confirm_keyboard(),
                parse_mode='HTML',
            )
            await crud.set_confirm_msg_id(team.telegram_id, msg.message_id)
            sent += 1
        except Exception:
            continue
    return sent


async def confirm_participation(telegram_id: int, bot: Bot, chat_id: int, msg_id: int) -> bool:
    """Подтверждает + удаляет сообщение с кнопками."""
    try:
        await bot.delete_message(chat_id=chat_id, message_id=msg_id)
    except Exception:
        pass
    return await crud.confirm_team(telegram_id)


async def decline_participation(telegram_id: int, bot: Bot, chat_id: int, msg_id: int) -> bool:
    """Отказ = удаление команды + удаление сообщения с кнопками."""
    try:
        await bot.delete_message(chat_id=chat_id, message_id=msg_id)
    except Exception:
        pass
    return await crud.delete_team(telegram_id)


async def broadcast(bot: Bot, text: str) -> tuple[int, int]:
    """Рассылает текст всем командам. Возвращает (sent, failed)."""
    teams = await crud.get_all_teams()
    sent = failed = 0
    for team in teams:
        try:
            await bot.send_message(team.telegram_id, text, parse_mode='HTML')
            sent += 1
        except Exception:
            failed += 1
    return sent, failed


def build_excel(teams) -> bytes:
    """Собирает Excel-файл из списка команд, возвращает bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = 'Команды'

    headers = [
        '#',
        'Команда',
        'Капитан',
        'Игрок 2',
        'Игрок 3',
        'Запасной',
        'Телефон',
        'Telegram',
        'Подтверждено',
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F6AA5')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, t in enumerate(teams, 2):
        ws.append(
            [
                row - 1,
                t.team_name,
                t.captain_name,
                t.player2,
                t.player3,
                t.substitute,
                t.phone,
                f'@{t.tg_username}' if t.tg_username else '—',
                '✅ Да' if t.confirmed else '❌ Нет',
            ]
        )

    # авто-ширина колонок
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
